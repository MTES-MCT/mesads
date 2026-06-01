import logging
import re
from datetime import date, datetime

from django.db import transaction
from openpyxl import load_workbook

from mesads.app.models import ADSUpdateLog, ADSUser

logger = logging.getLogger(__name__)


CONDUCTEUR_PATTERN = re.compile(r"^Statut_conducteur_(\d+)$")


def as_boolean(value, none_authorized=True):
    if (value is None or value == "") and none_authorized:
        return None
    if value.lower() in ["oui", "yes", "o", "1"]:
        return True
    return False


def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def parse_statut(value):
    if value == "TITULAIRE":
        return ADSUser.TITULAIRE_EXPLOITANT
    if value == "REPRESENTANT":
        return ADSUser.LEGAL_REPRESENTATIVE
    if value == "SALARIE":
        return ADSUser.SALARIE
    if value == "COOPERATEUR":
        return ADSUser.COOPERATEUR
    if value == "LOCATAIRE_GERANT":
        return ADSUser.LOCATAIRE_GERANT
    return ""


@transaction.atomic()
def import_ads_from_excel(file, ads_manager, user):
    logger.info("Début de l'import")

    workbook = load_workbook(file, data_only=True)

    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    logger.info(f"{len(rows)} à traiter")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

    conducteur_indexes = []

    number_creation = 0
    number_update = 0

    for header in headers:
        match = CONDUCTEUR_PATTERN.match(header)
        if match:
            conducteur_indexes.append(match.group(1))

    for row_number, row in enumerate(rows[1:], start=2):
        row_dict = dict(zip(headers, row))
        number = row_dict.get("Numero")
        if number:
            data = {
                "ads_creation_date": as_date(row_dict.get("Date_creation")),
                "ads_in_use": as_boolean(
                    row_dict.get("Exploitation"), none_authorized=False
                ),
                "ads_renew_date": as_date(row_dict.get("Date_dernier_renouvellement")),
                "attribution_date": as_date(row_dict.get("Date_attribution")),
                "accepted_cpam": as_boolean(row_dict.get("CPAM")),
                "immatriculation_plate": row_dict.get("Immatriculation") or "",
                "vehicle_compatible_pmr": as_boolean(row_dict.get("PMR")),
                "eco_vehicle": as_boolean(row_dict.get("Eco_vehicule")),
                "owner_name": row_dict.get("Titulaire") or "",
                "owner_siret": row_dict.get("SIRET") or "",
                "owner_phone": row_dict.get("Telephone_fixe") or "",
                "owner_mobile": row_dict.get("Telephone_mobile") or "",
                "owner_email": row_dict.get("Email") or "",
            }

            ads, created = ads_manager.ads_set.update_or_create(
                number=number, defaults=data
            )

            if created:
                number_creation += 1
            else:
                number_update += 1

            ads.adsuser_set.all().delete()
            for index in conducteur_indexes:
                statut = parse_statut(row_dict.get(f"Statut_conducteur_{index}"))
                nom = row_dict.get(f"Nom_conducteur_{index}")
                siret = row_dict.get(f"SIRET_conducteur_{index}")
                carte_pro = row_dict.get(f"Carte_pro_conducteur_{index}")

                if statut or nom or siret or carte_pro:
                    ADSUser.objects.create(
                        ads=ads,
                        status=statut,
                        name=nom or "",
                        siret=siret or "",
                        license_number=carte_pro or "",
                    )
            ADSUpdateLog.create_for_ads(ads, user)

    logger.info(f"{number_creation} ADS créées")
    logger.info(f"{number_update} ADS mise à jours")
    logger.info("Fin de l'import")

    return
